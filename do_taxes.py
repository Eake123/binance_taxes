from jsonclass import Json
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import os
import logging
class get_all_trades:
    def __init__(self,trades=None) -> None:
        self.tax = trades
        self.wb = Workbook()
        self.tax_wb = self.wb.create_sheet('taxes')
        self.trade_history =  self.wb.create_sheet('trade_sheet')
    


    def pnl(self):
        v = 0
        all_trades,df = self.iter_trades()
        for trade in all_trades:
            trades = trade['trades']
            if len(trades) > 0:
                for i in trade['trades']:
                    v += list(i.values())[0]
        return v,df

    def create_df(self):
        df = self.tax[0][0][0]
        return {key:[] for key in df.keys()}
    def df_to_xlsm(self,df,sheet):
        rows = dataframe_to_rows(df)

        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)



    def single_pnl(self,l):
        p = self.populate_trades(l)
        s = 0
        for i in p:
            for key,value in i.items():
                s += value
        return s


    def create_sheets(self,df):
        self.df_to_xlsm(df,self.trade_history)
        used = []
        for sheetname in df['symbol']:
            if sheetname not in used:
                #MuTaBlE
                l = [x for x in df.iloc if x['symbol'] == sheetname]
                l2 = [x for x in df.iloc if x['symbol'] == sheetname]
                if len(l) > 0:
                    s = self.single_pnl(l2)
                    df_single = pd.DataFrame(l)
                    ws = self.wb.create_sheet(sheetname)
                    self.df_to_xlsm(df_single,ws)
                    row = len(df_single) + 5
                    ws.cell(row=row,column=1,value='total earned')
                    ws.cell(row=row,column=2,value=s)
            used.append(sheetname)


    def excel(self):
        total_earned,df = self.pnl()
        tax_rate = 0.5
        self.tax_wb.cell(row=1,column=1,value='total earned')
        self.tax_wb.cell(row=2,column=1,value='taxable amount')
        self.tax_wb.cell(row=3,column=1,value='tax_rate')

        self.tax_wb.cell(row=1,column=2,value=total_earned)
        self.tax_wb.cell(row=2,column=2,value=total_earned * tax_rate)
        self.tax_wb.cell(row=3,column=2,value=tax_rate)
        self.create_sheets(df)
        # self.df_to_xlsm(df)
        self.wb.save('trades_and_taxes.xlsx')

    def add_to_df(self,trade,df):
        for trades in trade:
            for key,value in trades.items():
                df[key].append(value)
        return df

    def iter_trades(self):
        all_trades = []
        df = self.create_df()
        for i in self.tax:
            for o in i:
                df = self.add_to_df(o,df)
                dic = self.FIFO(o)
                if len(dic['trades']) > 0:
                    all_trades.append(dic)
        return all_trades,pd.DataFrame(df)

    

    def FIFO(self,trades):
        dic = {
            'symbol':trades[0]['symbol'],
            'trades':self.populate_trades(trades)
        }
        return dic
        
    

    def populate_trades(self,trades):
        sold_list = self.buy_sell(trades,False)
        bought_list = self.buy_sell(trades,True)
        change_list = []
        for sold in sold_list:
            change,bought_list = self.match_order(sold,bought_list)
            change_list.append(change)
        return change_list



    def match_order(self,sold,bought_list):
        sold_qty = float(sold['qty'])
        sold_price = float(sold['price'])
        sold_commision = float(sold['commission'])
        sold_com_asset = sold['commissionAsset']
        value_change = 0
        done = False
        l = []
        for buy in bought_list:
            buy_qty = float(buy['qty'])
            buy_price = float(buy['price'])
            change_qty = sold_qty - buy_qty
            if change_qty < 0 and done == False:
                # value_change += sold_qty * sold_price - (sold_qty * buy_price)
                value_change += self.less_buy(sold_qty,sold_price,sold_commision,sold_com_asset,buy_price)
                buy['qty'] = buy_qty - sold_qty
                sold_qty = 0
                done = True
                l.append(buy)
            elif change_qty >= 0:
                # value_change += buy_qty * sold_price - (buy_qty * buy_price)
                value_change += self.less_buy(buy_qty,sold_price,sold_commision,sold_com_asset,buy_price)
                sold_qty -= buy_qty
            else:
                l.append(buy)
        return {sold['time']:value_change},l


    def less_buy(self,sold_qty,sold_price,sold_commision,sold_com_asset,buy_price):
        sold = sold_qty * sold_price - (sold_qty * buy_price)
        return sold
                    
    # def get_commision(self,sold_commision,sold_com_asset):
    #     if 'USD' in sold_com_asset:
    #         return sold_commision
    #     elif 'BNB' in sold_com_asset:
    #         return sold_commision * self.bnb
    #     else:
    #         sold_com_asset += 'USDT'
    #         return sold_commision * float(self.binance.get_avg_price(sold_com_asset)['price'])

    def buy_sell(self,trades,buyer):
        return [x for x in trades if x['isBuyer'] == buyer]
    






                    



if __name__ == '__main__':
    trade = Json('taxes.txt').readKey()
    g = get_all_trades(trade)
    g.excel()
    # print(g.iter_trades())